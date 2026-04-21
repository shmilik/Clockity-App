"""
add_bom_table_tool.py

Replaces generate_bom_excel with a new generate_bom_table tool that:
  1. Prints a formatted markdown table to the agent chat
  2. Auto-saves the Excel file as a background side-effect
  3. Returns both the table AND a download path so the user can export

Also updates the excel_writer system_message accordingly.
"""

import sqlite3, json

DB_PATH = "C:/Users/info/.autogenstudio/autogen04202.db"

# ── NEW TOOL SOURCE ────────────────────────────────────────────────────────────
NEW_TOOL_SOURCE = r'''
def generate_bom_table(bom_json: str, job_name: str = "Solar_Install") -> str:
    """Format the solar BOM as a visible markdown table for the agent chat,
    AND silently save a two-sheet Excel file (BOM + Order Sheet) to disk.

    The markdown table is the primary return value so it appears directly in the
    AutoGen Studio chat window.  The Excel file path is appended at the bottom
    so the user can open it immediately.

    Args:
        bom_json:  JSON string from calculate_solar_bom (must contain a "bom" list).
        job_name:  Job/site name used in the header and Excel filename.

    Returns:
        A markdown-formatted order table string (visible in chat) with the
        Excel file path appended.
    """
    import json, os, textwrap
    from datetime import datetime

    # ── Parse ─────────────────────────────────────────────────────────────────
    try:
        payload = json.loads(bom_json)
    except Exception:
        return "**Error:** bom_json is not valid JSON.\n\n```\n" + bom_json[:300] + "\n```"

    bom_list = payload.get("bom", payload)
    if isinstance(bom_list, dict):
        bom_list = [{"description": k, "part_number": "", "qty": v, "unit": "EA"}
                    for k, v in bom_list.items()]
    summary = payload.get("summary", "")

    # ── MFR lookup ────────────────────────────────────────────────────────────
    MFR_MAP = [
        ("XR-10-",       "IRIDG"),
        ("XR-100-",      "IRIDG"),
        ("QM-CF-",       "IRIDG"),
        ("QM-HUG",       "IRIDG"),
        ("QM-LM",        "IRIDG"),
        ("QM-QBP",       "IRIDG"),
        ("QM-QBB",       "IRIDG"),
        ("QM-TLF",       "IRIDG"),
        ("QM-LMST",      "IRIDG"),
        ("QM-JBX",       "IRIDG"),
        ("QMC-",         "IRIDG"),
        ("QMTR-",        "IRIDG"),
        ("QMUTM",        "IRIDG"),
        ("QMR-",         "IRIDG"),
        ("LFT-03",       "IRIDG"),
        ("29-4000",      "IRIDG"),
        ("70-020",       "IRIDG"),
        ("70-030",       "IRIDG"),
        ("GM-",          "IRIDG"),
        ("FF2-",         "IRIDG"),
        ("FRA-",         "IRIDG"),
        ("XR-LUG",       "IRIDG"),
        ("XR10-BOSS",    "IRIDG"),
        ("XR100-BOSS",   "IRIDG"),
        ("WEEB",         "IRIDG"),
        ("FAK",          "QCELL"),
        ("FAM",          "QCELL"),
        ("FBM",          "QCELL"),
        ("Q.PEAK",       "QCELL"),
        ("IQ7",          "ENP"),
        ("IQ8",          "ENP"),
        ("X-IQ",         "ENP"),
        ("X2-IQ",        "ENP"),
        ("XA-PLUG",      "ENP"),
        ("Q-CONN",       "ENP"),
        ("Q-TERM",       "ENP"),
        ("Q-SEAL",       "ENP"),
        ("Q-CLIP",       "ENP"),
        ("Q-LCF",        "ENP"),
        ("Q-12",         "ENP"),
        ("Q-BA",         "ENP"),
        ("ENV-IQ",       "ENP"),
        ("HMS-",         "HOYM"),
        ("HM-",          "HOYM"),
        ("SE-",          "SOLAD"),
        ("P401",         "SOLAD"),
        ("P505",         "SOLAD"),
        ("BR115",        "EATON"),
        ("BR120",        "EATON"),
        ("BR215",        "EATON"),
        ("BR220",        "EATON"),
        ("BR225",        "EATON"),
        ("BR230",        "EATON"),
        ("BR240",        "EATON"),
        ("BR250",        "EATON"),
        ("BR1515",       "EATON"),
        ("BR612",        "EATON"),
        ("BR816",        "EATON"),
        ("HOM",          "EATON"),
        ("ZIPTIE",       "TYTON"),
        ("QO",           "SQD"),
    ]

    def get_mfr(pn, desc):
        p = (pn or "").upper()
        d = (desc or "").upper()
        for prefix, code in MFR_MAP:
            if p.startswith(prefix.upper()) or prefix.upper() in p:
                return code
        if "IRONRIDGE" in d or "IRON RIDGE" in d: return "IRIDG"
        if "Q.CELLS" in d or "QCELL" in d:        return "QCELL"
        if "ENPHASE" in d:                          return "ENP"
        if "EATON" in d:                            return "EATON"
        if "HOYMILES" in d:                         return "HOYM"
        if "SOLAREDGE" in d:                        return "SOLAD"
        if "ZIP TIE" in d or "ZIPTIE" in d:        return "TYTON"
        return "MISC."

    # ── Enrich rows ───────────────────────────────────────────────────────────
    rows = []
    for item in bom_list:
        desc = item.get("description", "")
        pn   = item.get("part_number", "")
        qty  = item.get("qty", 0)
        unit = item.get("unit", "EA")
        mfr  = get_mfr(pn, desc)
        rows.append((mfr, pn, desc, qty, unit))

    # ── Build markdown table ──────────────────────────────────────────────────
    # Column widths (pad to content)
    w_seq  = 3
    w_mfr  = max(5,  max(len(r[0]) for r in rows))
    w_pn   = max(12, max(len(r[1]) for r in rows))
    w_desc = max(11, max(len(r[2]) for r in rows))
    w_qty  = max(3,  max(len(str(r[3])) for r in rows))
    w_unit = max(4,  max(len(r[4]) for r in rows))

    def sep(c="-"):
        return (
            "| " + c*w_seq + " | " + c*w_mfr + " | " + c*w_pn +
            " | " + c*w_desc + " | " + c*w_qty + " | " + c*w_unit + " |"
        )

    def row_line(seq, mfr, pn, desc, qty, unit):
        return (
            f"| {str(seq):<{w_seq}} "
            f"| {mfr:<{w_mfr}} "
            f"| {pn:<{w_pn}} "
            f"| {desc:<{w_desc}} "
            f"| {str(qty):>{w_qty}} "
            f"| {unit:<{w_unit}} |"
        )

    lines = []
    lines.append(f"## ORDER SHEET  --  {job_name.upper()}")
    lines.append(f"*Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}*")
    if summary:
        lines.append(f"> {summary}")
    lines.append("")
    lines.append(row_line("#", "MFR", "Catalog #", "Description", "Qty", "Unit"))
    lines.append(sep())
    for i, (mfr, pn, desc, qty, unit) in enumerate(rows, 1):
        lines.append(row_line(i, mfr, pn, desc, qty, unit))
    lines.append(sep())
    lines.append(f"| {'':>{w_seq}} | {'':>{w_mfr}} | {'':>{w_pn}} | "
                 f"{'**TOTAL: ' + str(len(rows)) + ' line items**':<{w_desc}} "
                 f"| {'':>{w_qty}} | {'':>{w_unit}} |")
    lines.append("")

    table_text = "\n".join(lines)

    # ── Save Excel (side-effect) ──────────────────────────────────────────────
    excel_path = ""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        output_dir = os.path.join(
            os.path.expanduser("~"),
            "OneDrive", "Desktop", "JobTracker", "instance", "bom_sheets"
        )
        os.makedirs(output_dir, exist_ok=True)

        DARK_BLUE  = "1F4E79"
        MID_BLUE   = "2E75B6"
        LIGHT_BLUE = "D6E4F0"
        LIGHT_GRAY = "F2F2F2"
        thin       = Side(style="thin")
        tborder    = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hc(cell, text, bg=DARK_BLUE, fg="FFFFFF", sz=11):
            cell.value     = text
            cell.font      = Font(bold=True, color=fg, size=sz)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = tborder

        def dc(cell, value, bg="FFFFFF", bold=False, ha="left"):
            cell.value     = value
            cell.font      = Font(bold=bold, color="000000")
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal=ha, vertical="center")
            cell.border    = tborder

        wb = openpyxl.Workbook()

        # -- Sheet 1: BOM (detailed) --
        ws1 = wb.active
        ws1.title = "BOM"
        ws1.merge_cells("A1:G1")
        ws1["A1"].value     = "SOLAR INSTALLATION -- BILL OF MATERIALS"
        ws1["A1"].font      = Font(bold=True, size=15, color=DARK_BLUE)
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 28

        ws1.merge_cells("A2:G2")
        ws1["A2"].value     = "Project: " + job_name + "    |    " + datetime.now().strftime("%B %d, %Y")
        ws1["A2"].font      = Font(italic=True, size=10)
        ws1["A2"].alignment = Alignment(horizontal="center")

        if summary:
            ws1.merge_cells("A3:G3")
            ws1["A3"].value     = summary
            ws1["A3"].font      = Font(italic=True, size=9)
            ws1["A3"].alignment = Alignment(horizontal="left")

        hdrs = ["#", "Description", "Part Number", "MFR", "Qty", "Unit", "Notes"]
        wids = [5,   44,            22,             8,     7,     7,      20]
        for col, h in enumerate(hdrs, 1):
            hc(ws1.cell(row=5, column=col), h)
        ws1.row_dimensions[5].height = 20

        r = 6
        for i, (mfr, pn, desc, qty, unit) in enumerate(rows, 1):
            bg = LIGHT_BLUE if i % 2 == 0 else "FFFFFF"
            dc(ws1.cell(row=r, column=1), i,    bg=bg, ha="center")
            dc(ws1.cell(row=r, column=2), desc, bg=bg)
            dc(ws1.cell(row=r, column=3), pn,   bg=bg)
            dc(ws1.cell(row=r, column=4), mfr,  bg=bg, ha="center")
            dc(ws1.cell(row=r, column=5), qty,  bg=bg, bold=True, ha="center")
            dc(ws1.cell(row=r, column=6), unit, bg=bg, ha="center")
            ws1.row_dimensions[r].height = 18
            r += 1
        for col, w in enumerate(wids, 1):
            ws1.column_dimensions[get_column_letter(col)].width = w
        ws1.freeze_panes = "A6"

        # -- Sheet 2: Order Sheet --
        ws2 = wb.create_sheet("Order Sheet")
        ws2.merge_cells("A1:E1")
        ws2["A1"].value     = job_name.upper() + "  --  SOLAR ORDER SHEET"
        ws2["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
        ws2["A1"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 26
        ws2.merge_cells("A2:E2")
        ws2["A2"].value     = "Generated: " + datetime.now().strftime("%B %d, %Y")
        ws2["A2"].font      = Font(italic=True, size=10)
        ws2["A2"].alignment = Alignment(horizontal="center")

        ord_hdrs = ["Seq", "MFR", "Catalog #", "Description", "Qty"]
        ord_wids = [6,     9,     22,           44,            7]
        for col, h in enumerate(ord_hdrs, 1):
            hc(ws2.cell(row=4, column=col), h, bg=MID_BLUE)
        ws2.row_dimensions[4].height = 20

        r = 5
        for i, (mfr, pn, desc, qty, unit) in enumerate(rows, 1):
            bg = LIGHT_GRAY if i % 2 == 0 else "FFFFFF"
            dc(ws2.cell(row=r, column=1), i,    bg=bg, ha="center")
            dc(ws2.cell(row=r, column=2), mfr,  bg=bg, bold=True, ha="center")
            dc(ws2.cell(row=r, column=3), pn,   bg=bg)
            dc(ws2.cell(row=r, column=4), desc, bg=bg)
            dc(ws2.cell(row=r, column=5), qty,  bg=bg, bold=True, ha="center")
            ws2.row_dimensions[r].height = 18
            r += 1
        # total row
        ws2.merge_cells(f"A{r}:D{r}")
        t = ws2[f"A{r}"]
        t.value     = f"TOTAL LINE ITEMS: {len(rows)}"
        t.font      = Font(bold=True, size=10, color="FFFFFF")
        t.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        t.alignment = Alignment(horizontal="right", vertical="center")
        ws2.row_dimensions[r].height = 20
        for col, w in enumerate(ord_wids, 1):
            ws2.column_dimensions[get_column_letter(col)].width = w
        ws2.freeze_panes = "A5"

        safe  = "".join(c if c.isalnum() or c in "-_ " else "_" for c in job_name)
        fname = "ORDER_" + safe + "_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        excel_path = os.path.join(output_dir, fname)
        wb.save(excel_path)

    except Exception as ex:
        excel_path = f"(Excel save failed: {ex})"

    # ── Final return: table + file path ──────────────────────────────────────
    footer = f"\n---\n**Excel file saved:** `{excel_path}`  \n*(Open in Excel to view the formatted BOM and Order Sheet tabs)*"
    return table_text + footer
'''

# ── Validate ──────────────────────────────────────────────────────────────────
try:
    compile(NEW_TOOL_SOURCE.strip(), "<generate_bom_table>", "exec")
    print("Source compiles OK")
except SyntaxError as e:
    print(f"SYNTAX ERROR: {e}")
    raise

# ── New system message ─────────────────────────────────────────────────────────
NEW_SYSTEM_MSG = (
    "You are the Order Sheet generator for solar installations. "
    "When given BOM JSON from the BOM Calculator, call generate_bom_table with "
    "the BOM JSON and the job name from the engineering sheet (use 'Solar_Install' "
    "if unknown). The tool will output a formatted order table directly in this chat "
    "AND automatically save an Excel file. After the tool runs, tell the user the Excel "
    "file path so they can open it. Do not add extra commentary beyond the table and "
    "the file path. Then say TERMINATE."
)

# ── Patch the DB ──────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
cur  = conn.cursor()
cur.execute("SELECT component FROM team WHERE id=4")
data = json.loads(cur.fetchone()[0])

patched_tool = False
patched_msg  = False

for p in data["config"]["participants"]:
    if p["config"]["name"] != "excel_writer":
        continue

    # Update system message
    p["config"]["system_message"] = NEW_SYSTEM_MSG
    patched_msg = True
    print("Updated excel_writer system_message")

    tools = p["config"].get("workbench", {}).get("config", {}).get("tools", [])

    # Replace generate_bom_excel with generate_bom_table
    removed = [t for t in tools if t["config"]["name"] == "generate_bom_excel"]
    kept    = [t for t in tools if t["config"]["name"] != "generate_bom_excel"]

    if removed:
        # Re-use the existing tool object structure, just swap name + source
        new_tool = removed[0]
        new_tool["config"]["name"]        = "generate_bom_table"
        new_tool["config"]["description"] = (
            "Format the BOM as a markdown table visible in the agent chat and "
            "save an Excel file (BOM + Order Sheet tabs) to disk."
        )
        new_tool["config"]["source_code"] = NEW_TOOL_SOURCE.strip()
        kept.append(new_tool)
        patched_tool = True
        print("Replaced generate_bom_excel → generate_bom_table")
    else:
        # No existing tool found — nothing to replace
        print("WARNING: generate_bom_excel not found; tool NOT added")

    p["config"]["workbench"]["config"]["tools"] = kept

if not patched_tool:
    print("ERROR: tool was not patched")
else:
    cur.execute("UPDATE team SET component=? WHERE id=4", (json.dumps(data),))
    conn.commit()
    print("Saved to DB (team id=4)")

conn.close()

# ── Functional test ───────────────────────────────────────────────────────────
print("\nRunning functional test...")

import sqlite3 as _sq, json as _js
_conn = _sq.connect(DB_PATH)
_cur  = _conn.cursor()
_cur.execute("SELECT component FROM team WHERE id=4")
_data = _js.loads(_cur.fetchone()[0])
_conn.close()

_src = None
for p in _data["config"]["participants"]:
    if p["config"]["name"] == "excel_writer":
        for t in p["config"]["workbench"]["config"]["tools"]:
            if t["config"]["name"] == "generate_bom_table":
                _src = t["config"]["source_code"]

if not _src:
    print("ERROR: generate_bom_table not found in DB after patch")
else:
    ns = {}
    exec(_src, ns)
    fn = ns["generate_bom_table"]

    test_bom = _js.dumps({
        "summary": "27 panels | Q.CELLS Q.MI | IronRidge XR10 | Comp shingle",
        "bom": [
            {"description": "Q.CELLS Q.PEAK DUO BLK ML-G10+ 430W Panel", "part_number": "Q.PEAK-DUO-BLK-430", "qty": 27, "unit": "EA"},
            {"description": "IronRidge XR10 168in Rail Stick", "part_number": "XR-10-168M", "qty": 21, "unit": "EA"},
            {"description": "IronRidge XR10 Splice Kit",       "part_number": "XR-10-SP",   "qty": 10, "unit": "EA"},
            {"description": "IronRidge Comp Shingle L-Foot",   "part_number": "LFT-03-B1",  "qty": 90, "unit": "EA"},
            {"description": "IronRidge XR10 T-Bolt",           "part_number": "XR10-BOSS-01-M1", "qty": 90, "unit": "EA"},
            {"description": "Lag Screw 5/16in x 3in",          "part_number": "LAG-516-3",  "qty": 270, "unit": "EA"},
            {"description": "IronRidge Mid Clamp",             "part_number": "29-4000-077","qty": 48,  "unit": "EA"},
            {"description": "IronRidge End Clamp",             "part_number": "29-4000-077-END","qty": 12, "unit": "EA"},
            {"description": "IronRidge Ground Lug",            "part_number": "XR-LUG-03-A1","qty": 3,  "unit": "EA"},
            {"description": "10in Zip Tie Pack (100-ct)",      "part_number": "ZIPTIE-10IN-PK","qty": 1, "unit": "PK"},
            {"description": "Eaton BR 15A/2P Combiner Breaker","part_number": "BR1515",     "qty": 1,  "unit": "EA"},
            {"description": "Eaton BR 20A/2P Branch Breaker",  "part_number": "BR220",      "qty": 3,  "unit": "EA"},
            {"description": "Q.CELLS AC Connector Female",     "part_number": "FAK440E8G-BB","qty": 27, "unit": "EA"},
            {"description": "Q.CELLS AC Connector Male",       "part_number": "FAM365E7G-BB","qty": 27, "unit": "EA"},
        ]
    })

    result = fn(test_bom, "Jeffrey Hanson")
    print(result)
