"""
add_order_sheet.py
Rewrites the generate_bom_excel tool in the Solar BOM Team (id=4) DB.

Changes:
  1. Fixes BOM parsing (was wrongly calling .items() on a list)
  2. Adds MFR vendor code lookup for every part
  3. Sheet 1 "BOM" -- detailed reference with all fields
  4. Sheet 2 "Order Sheet" -- clean Sequence / MFR / Catalog # / Description / Qty
     matching the Jeffrey Hanson ORDER SHEET format
"""

import sqlite3, json

DB_PATH = "C:/Users/info/.autogenstudio/autogen04202.db"

NEW_SOURCE = r'''
def generate_bom_excel(bom_json: str, job_name: str = "Solar_Install") -> str:
    """Generate a formatted BOM Excel with two sheets:
      - "BOM"         -- full detail (description, part number, MFR, qty, unit)
      - "Order Sheet" -- clean Sequence / MFR / Catalog # / Description / Qty
                         matching the standard Solar Order Sheet format

    Args:
        bom_json:  JSON string from calculate_solar_bom (must contain a "bom" list).
        job_name:  Job/site name used in the filename and sheet headers.

    Returns:
        Path to the saved Excel file.
    """
    import json, os
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Output directory ──────────────────────────────────────────────────────
    output_dir = os.path.join(
        os.path.expanduser("~"),
        "OneDrive", "Desktop", "JobTracker", "instance", "bom_sheets"
    )
    os.makedirs(output_dir, exist_ok=True)

    # ── Parse BOM JSON ────────────────────────────────────────────────────────
    try:
        payload = json.loads(bom_json)
    except Exception:
        return "Error: bom_json is not valid JSON. Got: " + bom_json[:200]

    bom_list = payload.get("bom", payload)
    if isinstance(bom_list, dict):          # old dict fallback
        bom_list = [{"description": k, "part_number": "", "qty": v, "unit": "EA"}
                    for k, v in bom_list.items()]
    summary = payload.get("summary", "")

    # ── MFR vendor-code lookup ────────────────────────────────────────────────
    MFR_MAP = [
        # (prefix/substring to match in part_number, MFR code)
        # IronRidge
        ("XR-10-",       "IRIDG"),
        ("XR-100-",      "IRIDG"),
        ("QM-CF-",       "IRIDG"),
        ("QM-HUG",       "IRIDG"),
        ("QM-LM",        "IRIDG"),
        ("QM-QBP",       "IRIDG"),
        ("QM-QBB",       "IRIDG"),
        ("QM-TLF",       "IRIDG"),
        ("QM-LMST",      "IRIDG"),
        ("QM-JBX",       "IRIDG"),
        ("QMC-",         "IRIDG"),
        ("QMTR-",        "IRIDG"),
        ("QMUTM",        "IRIDG"),
        ("QMCMT",        "IRIDG"),
        ("QMCPC",        "IRIDG"),
        ("QMCPT",        "IRIDG"),
        ("QMCC-",        "IRIDG"),
        ("QMR-",         "IRIDG"),
        ("XR-LUG",       "IRIDG"),
        ("XR10-BOSS",    "IRIDG"),
        ("XR100-BOSS",   "IRIDG"),
        ("XR1000-BOSS",  "IRIDG"),
        ("LFT-03",       "IRIDG"),
        ("29-4000",      "IRIDG"),
        ("70-020",       "IRIDG"),
        ("70-030",       "IRIDG"),
        ("GM-",          "IRIDG"),
        ("FF2-",         "IRIDG"),
        ("FV-",          "IRIDG"),
        ("FRA-",         "IRIDG"),
        # Q.CELLS
        ("FAK",          "QCELL"),
        ("FAM",          "QCELL"),
        ("FBM",          "QCELL"),
        ("Q.PEAK",       "QCELL"),
        ("QCELL",        "QCELL"),
        # Enphase
        ("IQ7",          "ENP"),
        ("IQ8",          "ENP"),
        ("X-IQ",         "ENP"),
        ("X2-IQ",        "ENP"),
        ("XA-PLUG",      "ENP"),
        ("Q-CONN",       "ENP"),
        ("Q-TERM",       "ENP"),
        ("Q-SEAL",       "ENP"),
        ("Q-CLIP",       "ENP"),
        ("Q-LCF",        "ENP"),
        ("Q-12",         "ENP"),
        ("Q-BA",         "ENP"),
        ("ENV-IQ",       "ENP"),
        # Eaton
        ("BR115",        "EATON"),
        ("BR120",        "EATON"),
        ("BR215",        "EATON"),
        ("BR220",        "EATON"),
        ("BR225",        "EATON"),
        ("BR230",        "EATON"),
        ("BR240",        "EATON"),
        ("BR250",        "EATON"),
        ("BR1515",       "EATON"),
        ("BR612",        "EATON"),
        ("BR816",        "EATON"),
        ("HOM",          "EATON"),
        # Hoymiles
        ("HMS-",         "HOYM"),
        ("HM-",          "HOYM"),
        # SolarEdge
        ("SE-",          "SOLAD"),
        ("P401",         "SOLAD"),
        ("P505",         "SOLAD"),
        # Tyton / zip ties
        ("ZIPTIE",       "TYTON"),
        # SolarEdge optimizers
        ("P300",         "SOLAD"),
        ("P400",         "SOLAD"),
        ("P500",         "SOLAD"),
        # Schneider / Square D
        ("QO",           "SQD"),
        # NSI / wire nuts
        ("NSI",          "NSI"),
        # misc
        ("WEEB",         "IRIDG"),
    ]

    def get_mfr(part_number, description):
        pn = (part_number or "").upper()
        desc = (description or "").upper()
        for prefix, code in MFR_MAP:
            if pn.startswith(prefix.upper()) or prefix.upper() in pn:
                return code
        # fallback: check description
        if "IRONRIDGE" in desc or "IRON RIDGE" in desc:
            return "IRIDG"
        if "Q.CELLS" in desc or "QCELL" in desc:
            return "QCELL"
        if "ENPHASE" in desc:
            return "ENP"
        if "EATON" in desc:
            return "EATON"
        if "HOYMILES" in desc:
            return "HOYM"
        if "SOLAREDGE" in desc or "SOLAR EDGE" in desc:
            return "SOLAD"
        if "ZIP TIE" in desc or "ZIPTIE" in desc:
            return "TYTON"
        return "MISC."

    # ── Style helpers ─────────────────────────────────────────────────────────
    DARK_BLUE   = "1F4E79"
    MID_BLUE    = "2E75B6"
    LIGHT_BLUE  = "D6E4F0"
    DARK_GRAY   = "404040"
    LIGHT_GRAY  = "F2F2F2"
    thin_side   = Side(style="thin")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    def hdr_cell(cell, text, bg=DARK_BLUE, fg="FFFFFF", size=11, bold=True):
        cell.value     = text
        cell.font      = Font(bold=bold, color=fg, size=size)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    def data_cell(cell, value, bg="FFFFFF", bold=False, halign="left"):
        cell.value     = value
        cell.font      = Font(bold=bold, color="000000")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=halign, vertical="center")
        cell.border    = thin_border

    wb = openpyxl.Workbook()

    # =========================================================================
    # SHEET 1 -- BOM (detailed reference)
    # =========================================================================
    ws_bom = wb.active
    ws_bom.title = "BOM"

    # Title
    ws_bom.merge_cells("A1:G1")
    ws_bom["A1"].value     = "SOLAR INSTALLATION -- BILL OF MATERIALS"
    ws_bom["A1"].font      = Font(bold=True, size=16, color=DARK_BLUE)
    ws_bom["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_bom.row_dimensions[1].height = 30

    # Sub-title
    ws_bom.merge_cells("A2:G2")
    ws_bom["A2"].value     = (
        "Project: " + job_name +
        "    |    Generated: " + datetime.now().strftime("%B %d, %Y  %H:%M")
    )
    ws_bom["A2"].font      = Font(italic=True, size=10, color=DARK_GRAY)
    ws_bom["A2"].alignment = Alignment(horizontal="center")

    if summary:
        ws_bom.merge_cells("A3:G3")
        ws_bom["A3"].value     = summary
        ws_bom["A3"].font      = Font(italic=True, size=9)
        ws_bom["A3"].alignment = Alignment(horizontal="left")

    # Column headers
    bom_headers = ["#", "Description", "Part Number", "MFR", "Qty", "Unit", "Notes"]
    bom_widths   = [5,   44,            22,             8,     7,     7,      20]
    hr = 5
    for col, (h, _) in enumerate(zip(bom_headers, bom_widths), 1):
        hdr_cell(ws_bom.cell(row=hr, column=col), h)
    ws_bom.row_dimensions[hr].height = 20

    row = hr + 1
    for i, item in enumerate(bom_list, 1):
        desc    = item.get("description", "")
        pn      = item.get("part_number", "")
        qty     = item.get("qty", 0)
        unit    = item.get("unit", "EA")
        notes   = item.get("notes", "")
        mfr     = get_mfr(pn, desc)
        bg      = LIGHT_BLUE if i % 2 == 0 else "FFFFFF"

        data_cell(ws_bom.cell(row=row, column=1), i,     bg=bg, halign="center")
        data_cell(ws_bom.cell(row=row, column=2), desc,  bg=bg)
        data_cell(ws_bom.cell(row=row, column=3), pn,    bg=bg)
        data_cell(ws_bom.cell(row=row, column=4), mfr,   bg=bg, halign="center")
        data_cell(ws_bom.cell(row=row, column=5), qty,   bg=bg, bold=True, halign="center")
        data_cell(ws_bom.cell(row=row, column=6), unit,  bg=bg, halign="center")
        data_cell(ws_bom.cell(row=row, column=7), notes, bg=bg)
        ws_bom.row_dimensions[row].height = 18
        row += 1

    for col, width in enumerate(bom_widths, 1):
        ws_bom.column_dimensions[get_column_letter(col)].width = width

    ws_bom.freeze_panes = "A6"

    # =========================================================================
    # SHEET 2 -- Order Sheet (matches Jeffrey Hanson ORDER SHEET format)
    # =========================================================================
    ws_ord = wb.create_sheet("Order Sheet")

    # Title block
    ws_ord.merge_cells("A1:E1")
    ws_ord["A1"].value     = job_name.upper() + "  --  SOLAR ORDER SHEET"
    ws_ord["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws_ord["A1"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws_ord["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_ord.row_dimensions[1].height = 28

    ws_ord.merge_cells("A2:E2")
    ws_ord["A2"].value     = "Generated: " + datetime.now().strftime("%B %d, %Y")
    ws_ord["A2"].font      = Font(italic=True, size=10, color=DARK_GRAY)
    ws_ord["A2"].alignment = Alignment(horizontal="center")
    ws_ord.row_dimensions[2].height = 16

    # Column headers -- matching Jeffrey Hanson format
    ord_headers = ["Seq", "MFR", "Catalog #", "Description", "Qty"]
    ord_widths   = [6,     9,     22,           44,            7]
    hr2 = 4
    for col, h in enumerate(ord_headers, 1):
        hdr_cell(ws_ord.cell(row=hr2, column=col), h, bg=MID_BLUE)
    ws_ord.row_dimensions[hr2].height = 20

    row = hr2 + 1
    for i, item in enumerate(bom_list, 1):
        desc  = item.get("description", "")
        pn    = item.get("part_number", "")
        qty   = item.get("qty", 0)
        mfr   = get_mfr(pn, desc)
        bg    = LIGHT_GRAY if i % 2 == 0 else "FFFFFF"

        data_cell(ws_ord.cell(row=row, column=1), i,    bg=bg, halign="center")
        data_cell(ws_ord.cell(row=row, column=2), mfr,  bg=bg, halign="center", bold=True)
        data_cell(ws_ord.cell(row=row, column=3), pn,   bg=bg)
        data_cell(ws_ord.cell(row=row, column=4), desc, bg=bg)
        data_cell(ws_ord.cell(row=row, column=5), qty,  bg=bg, bold=True, halign="center")
        ws_ord.row_dimensions[row].height = 18
        row += 1

    # Totals row
    ws_ord.merge_cells(f"A{row}:D{row}")
    total_cell = ws_ord[f"A{row}"]
    total_cell.value     = "TOTAL LINE ITEMS: " + str(len(bom_list))
    total_cell.font      = Font(bold=True, size=10, color="FFFFFF")
    total_cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws_ord.row_dimensions[row].height = 20

    for col, width in enumerate(ord_widths, 1):
        ws_ord.column_dimensions[get_column_letter(col)].width = width

    ws_ord.freeze_panes = "A5"

    # ── Save ──────────────────────────────────────────────────────────────────
    safe   = "".join(c if c.isalnum() or c in "-_ " else "_" for c in job_name)
    fname  = "BOM_" + safe + "_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
    fpath  = os.path.join(output_dir, fname)
    wb.save(fpath)
    return "SUCCESS: BOM + Order Sheet Excel saved to: " + fpath
'''

# ── Validate the new source compiles cleanly ──────────────────────────────────
try:
    compile(NEW_SOURCE.strip(), "<generate_bom_excel>", "exec")
    print("Source compiles OK")
except SyntaxError as e:
    print(f"SYNTAX ERROR: {e}")
    raise

# ── Patch the DB ──────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
cur  = conn.cursor()
cur.execute("SELECT component FROM team WHERE id=4")
data = json.loads(cur.fetchone()[0])

patched = False
for p in data["config"]["participants"]:
    if p["config"]["name"] != "excel_writer":
        continue
    tools = p["config"].get("workbench", {}).get("config", {}).get("tools", [])
    for t in tools:
        if t["config"]["name"] == "generate_bom_excel":
            t["config"]["source_code"] = NEW_SOURCE.strip()
            patched = True
            print("Patched generate_bom_excel in excel_writer")

if not patched:
    print("WARNING: generate_bom_excel tool not found -- nothing patched")
else:
    cur.execute("UPDATE team SET component=? WHERE id=4", (json.dumps(data),))
    conn.commit()
    print("Saved to DB (team id=4)")

conn.close()

# ── Quick functional test ─────────────────────────────────────────────────────
print("\nRunning functional test...")
exec(NEW_SOURCE.strip(), {"__name__": "__test__"})

test_bom = json.dumps({
    "summary": "27 panels | Q.CELLS Q.MI | IronRidge XR10 | Comp shingle",
    "bom": [
        {"description": "Q.CELLS Q.PEAK DUO BLK ML-G10+ 430W Panel", "part_number": "Q.PEAK-DUO-BLK-430", "qty": 27, "unit": "EA"},
        {"description": "IronRidge XR10 168\" Rail Stick", "part_number": "XR-10-168M", "qty": 21, "unit": "EA"},
        {"description": "IronRidge XR10 Splice Kit", "part_number": "XR-10-SP", "qty": 10, "unit": "EA"},
        {"description": "IronRidge Comp Shingle L-Foot", "part_number": "LFT-03-B1", "qty": 90, "unit": "EA"},
        {"description": "IronRidge XR10 T-Bolt", "part_number": "XR10-BOSS-01-M1", "qty": 90, "unit": "EA"},
        {"description": "Lag Screw 5/16\" x 3\"", "part_number": "LAG-516-3", "qty": 270, "unit": "EA"},
        {"description": "IronRidge Mid Clamp", "part_number": "29-4000-077", "qty": 48, "unit": "EA"},
        {"description": "IronRidge End Clamp", "part_number": "29-4000-077-END", "qty": 12, "unit": "EA"},
        {"description": "IronRidge Ground Lug", "part_number": "XR-LUG-03-A1", "qty": 3, "unit": "EA"},
        {"description": "10\" Zip Tie Pack (100-ct)", "part_number": "ZIPTIE-10IN-PK", "qty": 1, "unit": "PK"},
        {"description": "Eaton BR 15A/2P Combiner Breaker", "part_number": "BR1515", "qty": 1, "unit": "EA"},
        {"description": "Eaton BR 20A/2P Branch Circuit Breaker", "part_number": "BR220", "qty": 3, "unit": "EA"},
        {"description": "Q.CELLS AC Connector (Female)", "part_number": "FAK440E8G-BB", "qty": 27, "unit": "EA"},
        {"description": "Q.CELLS AC Connector (Male)", "part_number": "FAM365E7G-BB", "qty": 27, "unit": "EA"},
        {"description": "Q.CELLS Terminator Cap", "part_number": "FAK-TERM-CAP", "qty": 1, "unit": "EA"},
    ]
})

result = generate_bom_excel(test_bom, "Jeffrey Hanson Test")
print(result)
